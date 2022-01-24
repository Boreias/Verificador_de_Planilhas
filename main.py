import openpyxl
import time


def cresce_colunas(lista_colunas):
    tamanho = len(lista_colunas[-1])
    if ord(lista_colunas[-1][-1]) < 90:
        lista_colunas.append(lista_colunas[-1][0:-1] + chr(ord(lista_colunas[-1][-1]) + 1))
    elif tamanho > 1 and ord(lista_colunas[-1][-2]) < 90:
        lista_colunas.append(lista_colunas[-1][0:-2] + chr(ord(lista_colunas[-1][-2]) + 1) + 'A')
    elif tamanho > 2 and ord(lista_colunas[-1][-3]) < 90:
        lista_colunas.append(lista_colunas[-1][0:-3] + chr(ord(lista_colunas[-1][-3]) + 1) + 'AA')
    elif tamanho == 1:
        lista_colunas.append('AA')
    elif tamanho == 2:
        lista_colunas.append('AAA')
    else:
        print('Por favor, contate o suporte')
        return
    return lista_colunas


def descobre_numero(coluna):
    lista_colunas = ['A']
    while lista_colunas[-1] != coluna:
        cresce_colunas(lista_colunas)
    return len(lista_colunas)


def descobre_colunas(quantidade):
    lista_colunas = ['A']
    while len(lista_colunas) < quantidade:
        lista_colunas = cresce_colunas(lista_colunas)
    return lista_colunas


print('Iniciando Programa')

while True:
    try:
        planilha1 = input('Digite o nome da planilha mais atual: ')
        planilha1 = openpyxl.load_workbook(filename=planilha1+ '.xlsx')
        aba1 = planilha1.active
        break
    except:
        print('Planilha nao encontrada, favor digitar novamente: ')

linhas_aba1 = aba1.max_row
colunas_aba1 = aba1.max_column

while True:
    try:
        planilha2 = input('Digite o nome da planilha desatualizada: ')
        planilha2 = openpyxl.load_workbook(filename=planilha2 + '.xlsx')
        aba2 = planilha2.active
        break
    except:
        print('Planilha nao encontrada, favor digitar novamente: ')

linhas_aba2 = aba2.max_row
colunas_aba2 = aba2.max_column

coluna = input('Digite a coluna a ser analisada: ').upper()
coluna = descobre_numero(coluna)

while True:
    inicio = int(input('As planilhas apresentam cabecalho?\n1 - Nao\n2 - Sim\n'))
    if inicio == 1 or inicio == 2:
        break
    else:
        print('Por favor, digite o numero referente a resposta')

nome_planilha = input('Digite o nome da planilha a ser salva: ')

planilha3 = openpyxl.Workbook()
aba3 = planilha3.active

lista_colunas = descobre_colunas(colunas_aba1 + 1)

if inicio == 1:
    contador = 1
    
else:
    contador = 2
    for i in range(1, colunas_aba1 + 1):
        aba3.cell(1, i, value=aba1.cell(1, i).value)
    aba3.cell(1, colunas_aba1 + 1, value='Status')

for i in range(inicio, linhas_aba1 + 1):
    valor_analise = aba1.cell(i, coluna).value
    k = inicio
    while k <= linhas_aba2:
        if valor_analise == aba2.cell(k, coluna).value:
            for j in range(1, colunas_aba1 + 1):
                if aba1.cell(i, j).value != aba2.cell(k, j).value:
                    for h in range(1, colunas_aba1 + 1):
                        aba3.cell(contador, h, value=aba1.cell(i, h).value)
                        aba3.cell(contador + 1, h, value=aba2.cell(k, h).value)
                        if aba1.cell(i, h).value != aba2.cell(k, h).value:
                            aba3.cell(contador, h).font = openpyxl.styles.Font(color="00FF0000")
                            aba3.cell(contador, h).fill = openpyxl.styles.PatternFill(fill_type='solid', start_color='FFD890', end_color='FFD890')
                            aba3.cell(contador + 1, h).font = openpyxl.styles.Font(color="000000FF")
                            aba3.cell(contador + 1, h).fill = openpyxl.styles.PatternFill(fill_type='solid', start_color='ADD8E6', end_color='ADD8E6')
                    else:
                        aba3.cell(contador, colunas_aba1 + 1, value='Desatualizado')
                        aba3.cell(contador + 1, colunas_aba1 + 1, value='Atualizado')
                        contador += 2
                    break
            aba2.delete_rows(k)
            linhas_aba2 -= 1
            k -= 1
            break
        k += 1
    else:
        for h in range(1, colunas_aba1 + 1):
            aba3.cell(contador, h, value=aba1.cell(i, h).value)
        else:
            aba3.cell(contador, colunas_aba1 + 1, value='Adicionado')
            contador += 1

linhas_aba2 = aba2.max_row

for i in range(inicio, linhas_aba2 + 1):
    for j in range(1, colunas_aba2 + 1):
        aba3.cell(contador, j, value=aba2.cell(i, j).value)
    else:
        aba3.cell(contador, colunas_aba1 + 1, value='Excluido')
        contador += 1

planilha3.save(nome_planilha + '.xlsx')

print('Programa Finalizado')

time.sleep(3.5)
